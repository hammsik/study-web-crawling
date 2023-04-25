from selenium import webdriver 
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys 
from selenium.webdriver.common.by import By
from openpyxl import *
import time 
import traceback

def loadAllContents(): 
    prev_contents_len = 0 
    while True: 
        time.sleep(2)
        closeAd()
        contents = driver.find_elements(By.TAG_NAME, 'ytd-video-renderer') 
        print("현재 로드된 영상 개수:", len(contents)) 
        if prev_contents_len == len(contents) or len(contents) > 300: # 300개 이상이면 컷
            break 
        prev_contents_len = len(contents) 
        html.send_keys(Keys.END) 
    return contents

def pickVideo(sheet):
    cnt = 1
    for c in contents:
        print(cnt, "번째 video")
        time.sleep(1)

        closeAd()

        # 동영상 클릭
        toClick = c.find_element(By.TAG_NAME, "img")
        action.move_to_element(toClick).click().perform()

        time.sleep(2)
        html.send_keys(Keys.END)
        time.sleep(1)
        html.send_keys(Keys.PAGE_DOWN)

        time.sleep(0.3)
        html.send_keys(Keys.SPACE)
        time.sleep(0.2)
        html.send_keys(Keys.HOME)
        time.sleep(0.5)

        title = driver.find_element(By.XPATH, '//*[@id="title"]/h1/yt-formatted-string').text
        print("영상제목:", title)
        time.sleep(0.5)
        try:
            header = html.find_element(By.XPATH, '//*[@id="count"]/yt-formatted-string/span[2]')
            print("댓글 수:", header.text)
            loadAllComments()
            print("댓글 로드완료")
            crawling(sheet, title)
        except:
            print(traceback.format_exc())
        print()

        driver.back()
        cnt += 1

def loadAllComments():
    prev_comment_len = 0
    time.sleep(2)
    while True:

        closeAd()

        time.sleep(1)
        comments = html.find_elements(By.XPATH, '//*[@id="contents"]/ytd-comment-thread-renderer')
        print("현재 로드 된 댓글 개수:", prev_comment_len, len(comments))
        if prev_comment_len == len(comments) or len(comments) > 500: # 댓글은 500개 까지만
            print("로드완료:", prev_comment_len, len(comments))
            break
        prev_comment_len = len(comments)
        action.move_to_element(comments[-1]).perform()
        time.sleep(0.5)
        html.send_keys(Keys.PAGE_DOWN)
    loadAllReplies()

def loadAllReplies():
    time.sleep(1)
    replies = html.find_elements(By.XPATH, '//*[@id="more-replies"]/yt-button-shape/button/yt-touch-feedback-shape/div/div[2]')
    print("대댓이 달린 댓글 개수:", len(replies))
    time.sleep(0.5)
    clickReplies(replies)

def clickReplies(replies):
    cnt = 1
    html.send_keys(Keys.HOME)
    time.sleep(0.5)
    for r in replies:
        action.move_to_element(r).perform()
        time.sleep(1.5)
        closeAd()
        r.click()

        if cnt % 10 == 0:
            tmp = html.find_elements(By.ID, 'content-text')
            time.sleep(3)
            if len(tmp) > 1000:
                break

        print("클릭:", cnt)
        cnt += 1
        time.sleep(1)
        html.send_keys(Keys.PAGE_DOWN)
        time.sleep(0.5)
        closeAd()
        time.sleep(0.5)
        html.send_keys(Keys.PAGE_UP)
        showMore = html.find_elements(By.XPATH, '/html/body/ytd-app/div[1]/ytd-page-manager/ytd-watch-flexy/div[5]/div[1]/div/div[2]/ytd-comments/ytd-item-section-renderer/div[3]/ytd-comment-thread-renderer/div/ytd-comment-replies-renderer/div[1]/div[2]/div[1]/ytd-continuation-item-renderer/div[2]/ytd-button-renderer/yt-button-shape/button/yt-touch-feedback-shape/div/div[2]')
        time.sleep(0.5)
        print(len(showMore))

        try:
            while len(showMore) > 0:
                time.sleep(1)
                print("대댓글 진입:", len(showMore))
                action.move_to_element(showMore[0]).perform()
                time.sleep(1)
                closeAd()
                showMore[0].click()
                html.send_keys(Keys.PAGE_DOWN)
                time.sleep(1)
                html.send_keys(Keys.PAGE_UP)
        except:
            print("대댓글 로딩 완료")
            
def crawling(sheet, title):
    commentsBox = html.find_elements(By.ID, 'content-text')
    time.sleep(1)
    print(len(commentsBox))
    for comment in commentsBox:
        sheet.append([title, comment.text])
    time.sleep(5)
    excelfile.save('./crawlingText.xlsx')
    time.sleep(5)

def closeAd(): # 유튜브프리미엄 광고 팝업 닫기
    try:    
        html.find_element(By.XPATH, '//*[@id="dismiss-button"]/yt-button-shape/button/yt-touch-feedback-shape/div/div[2]').click()
    except:
        pass

if  __name__  ==  "__main__" :
    keyWordList = [
        '뉴스',
        '예능',
        '게임',
        '여가',
        '학습',
        '여행',
        '인기급상승',
        'BODA 보다',
        '비디오머그',
        '패션',
        '영화리뷰'
    ]

    for keyWord in keyWordList:
        search_key = keyWord # 검색할 단어 

        excelfile = load_workbook('crawlingText.xlsx') # test.xlsx 파일을 load 
        excelfile.create_sheet(index=3, title=search_key) # sheet를 검색할 단어 이름으로 생성 
        sheet = excelfile[search_key] # 생성한 시트를 변수에 할당 
        ti = sheet.cell(row=1, column=1) # 시트의 1행, 1열에 "영상 제목"이라는 문자열 삽입 
        ti.value = "영상 제목" 

        comment_list = sheet.cell(row=1, column=2) # 시트의 1행, 2열에 "댓글"이라는 문자열 삽입 
        comment_list.value = "댓글" 
        excelfile.save('./crawlingText.xlsx') # 파일 저장

        url = 'https://www.youtube.com/results?search_query={}'.format(search_key)
        driver = webdriver.Chrome('C:\WooooooooooW\etc\chromedriver_win32\chromedriver')
        driver.get(url)
        options = webdriver.ChromeOptions()
        options.add_argument("disable-gpu") # gpu 사용량 최소화
        html = driver.find_element(By.TAG_NAME, "html")

        #원하는 영상으로 스크롤 후 클릭하기 위한 ActionChains 객체 생성
        action = ActionChains(driver)

        contents = loadAllContents()
        print("최종 로드된 영상 개수:", len(contents))
        pickVideo(sheet)