from openpyxl import * 

search_key = "메이플스토리" # 검색할 단어 

excelfile = load_workbook('test.xlsx') # test.xlsx 파일을 load 
excelfile.create_sheet(index=3, title=search_key) # sheet를 검색할 단어 이름으로 생성 
sheet = excelfile[search_key] # 생성한 시트를 변수에 할당 
ti = sheet.cell(row=1, column=1) # 시트의 1행, 1열에 "영상 제목"이라는 문자열 삽입 
ti.value = "영상 제목" 

comment_list = sheet.cell(row=1, column=2) # 시트의 1행, 2열에 "댓글"이라는 문자열 삽입 
comment_list.value = "댓글" 
excelfile.save('./test.xlsx') # 파일 저장